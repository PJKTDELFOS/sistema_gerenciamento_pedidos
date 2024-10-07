import sqlite3
from openpyxl import load_workbook
from datetime import datetime
import os
def criar_pedido(name, n_pedido, data_origin, cnpj, contratante, endereco_contratante,
                 n_contrato, Ata_RP, empenho, o_f, data_recebimento_empenho, nf, contato,
                 telefone, email, objeto, data_hora_entrega, local_entrega, U_F, qtde, obs,
                 coord):
    template_form = 'modelo_pedido.xlsx'
    workbook = load_workbook(filename=template_form)
    worksheet = workbook['sheet']
    worksheet['I9'] = n_pedido
    worksheet['I10'] = data_origin
    worksheet['I12'] = cnpj
    worksheet['B12'] = contratante
    worksheet['B13'] = endereco_contratante
    worksheet['A15'] = n_contrato
    worksheet['B15'] = Ata_RP
    worksheet['C15'] = empenho
    worksheet['D15'] = o_f
    worksheet['E15'] = data_recebimento_empenho
    worksheet['F15'] = nf
    worksheet['G15'] = contato
    worksheet['H15'] = telefone
    worksheet['I15'] = email
    worksheet['D17'] = objeto
    worksheet['B16'] = data_hora_entrega
    worksheet['D16'] = local_entrega
    worksheet['A21'] = U_F
    worksheet['B21'] = qtde
    worksheet['A23'] = obs
    worksheet['B51'] = coord
    status_pedido=input('status do pedido')
    valor=float(input("valor do Pedido:"))
    inserir_pedido_banco_dados(n_pedido, n_contrato, empenho, objeto, data_hora_entrega, coord, data_origin,status_pedido,valor)
    workbook.save(filename=f'{name}.xlsx')
    print(f"Planilha salva como '{name}.xlsx'.")
def atualizar_pedido(name, newn_pedido, data_att, id, newcnpj=None, newcontratante=None,
                     newendereco_contratante=None, newn_contrato=None, newAta_RP=None,
                     newempenho=None, newo_f=None, newdata_recebimento_empenho=None, nf=None,
                     newcontato=None, newtelefone=None, newemail=None, newobjeto=None,
                     newdata_hora_entrega=None, newlocal_entrega=None, newU_F=None,
                     newqtde=None, newobs=None, newcoord=None):
    filename = f'{name}.xlsx'
    workbook = load_workbook(filename=filename)
    worksheet = workbook['sheet']
    worksheet['I09'] = newn_pedido or worksheet['I09'].value
    worksheet['I11'] = data_att
    worksheet['I12'] = newcnpj or worksheet['I12'].value
    worksheet['B12'] = newcontratante or worksheet['B12'].value
    worksheet['B13'] = newendereco_contratante or worksheet['B13'].value
    worksheet['A15'] = newn_contrato or worksheet['A15'].value
    worksheet['B15'] = newAta_RP or worksheet['B15'].value
    worksheet['C15'] = newempenho or worksheet['C15'].value
    worksheet['D15'] = newo_f or worksheet['D15'].value
    worksheet['E15'] = newdata_recebimento_empenho or worksheet['E15'].value
    worksheet['F15'] = nf or worksheet['F15'].value
    worksheet['G15'] = newcontato or worksheet['G15'].value
    worksheet['H15'] = newtelefone or worksheet['H15'].value
    worksheet['I15'] = newemail or worksheet['I15'].value
    worksheet['D17'] = newobjeto or worksheet['D17'].value
    worksheet['B16'] = newdata_hora_entrega or worksheet['B16'].value
    worksheet['D16'] = newlocal_entrega or worksheet['D16'].value
    worksheet['A21'] = newU_F or worksheet['A21'].value
    worksheet['B21'] = newqtde or worksheet['B21'].value
    worksheet['A23'] = newobs or worksheet['A23'].value
    worksheet['B51'] = newcoord or worksheet['B51'].value
    novo_status=input('atualize o status do pedido')
    novo_valor=float(input("Valor atualizado do Pedido"))
    att_pedido_DB(id, newn_pedido, newn_contrato, newempenho, newobjeto,
                  newdata_hora_entrega, newcoord, data_att,novo_status,novo_valor)

    workbook.save(filename=filename)
    print(f"Pedido '{name}.xlsx' atualizado com sucesso.")
def conect():
    return sqlite3.connect("pedidos.db")
def criar_tabela():
    conn = conect()
    cursor = conn.cursor()
    cursor.execute('''
          create table if not exists pedidos (
          id integer primary key autoincrement,
          numero text not null,
          contrato text not null,
          empenho text not null,
          objeto text not null,
          data_entrega text not null,
          coordenador text not null,
          data_criacao_att text not null,
          status text not null,
          valor float not null          
          )
          ''')
    conn.commit()
    conn.close()
def inserir_pedido_banco_dados(n_pedido, contrato, empenho,
                               objeto, data_hora_entrega,
                               coordenador, data_origin,status_pedido,valor):
    conn = conect()
    cursor = conn.cursor()
    cursor.execute(
        'insert into pedidos ('
        'numero,'
        'contrato,'
        'empenho,'
        'objeto,'
        'data_entrega,'
        'coordenador,'
        'data_criacao_att,'
        'status,'
        'valor)'
        'values (?,?,?,?,?,?,?,?,?)',
        (n_pedido, contrato,
         empenho, objeto, data_hora_entrega,
         coordenador, data_origin,status_pedido,valor)
    )
    conn.commit()
    conn.close()
def att_pedido_DB(id, newn_pedido=None,
                  newn_contrato=None, newempenho=None,
                  newobjeto=None,newdata_hora_entrega=None,
                  newcoord=None, data_att=None,novo_status=None,novo_valor=None):
    conn = conect()
    cursor = conn.cursor()
    update_columns = []
    parameters = []
    if newn_pedido is not None:
        update_columns.append("numero = ?")
        parameters.append(newn_pedido)
    if newn_contrato is not None:
        update_columns.append("contrato = ?")
        parameters.append(newn_contrato)
    if newempenho is not None:
        update_columns.append("empenho = ?")
        parameters.append(newempenho)
    if newobjeto is not None:
        update_columns.append("objeto = ?")
        parameters.append(newobjeto)
    if newdata_hora_entrega is not None:
        update_columns.append("data_entrega = ?")
        parameters.append(newdata_hora_entrega)
    if newcoord is not None:
        update_columns.append("coordenador = ?")
        parameters.append(newcoord)
    if data_att is not None:
        update_columns.append("data_criacao_att = ?")
        parameters.append(data_att)
    if novo_status is not None:
        update_columns.append("status_pedido=?")
        parameters.append(novo_status)
    if novo_valor is not None:
        update_columns.append("valor=?")
        parameters.append(novo_valor)
    sql = f'''
        UPDATE pedidos
        SET {', '.join(update_columns)}
        WHERE id = ?
    '''
    parameters.append(id)
    cursor.execute(sql, parameters)
    conn.commit()
    conn.close()
def ver_dados():
    while True:
        todos = input('digite o que deseja,[A] para visualização geral ou [B] pára ver registro unico, [C] para filtrar :').strip().upper()
        if todos == 'A':
            conn = conect()
            cursor = conn.cursor()
            cursor.execute("select * from pedidos")
            resultado = cursor.fetchall()
            conn.close()
            return resultado
        elif todos == 'B':
            indice = int(input('indique o id que deseja buscar:'))
            conn = conect()
            cursor = conn.cursor()
            cursor.execute("select * from pedidos where id =?", (indice,))  # sempre se
            # lembrar da maldita virgula
            usuario = cursor.fetchone()
            conn.close()
            return usuario
def filtro_db(filtro, filtrado):
    conn = conect()
    cursor = conn.cursor()
    busca = f'SELECT * FROM pedidos WHERE {filtro} = ?'
    cursor.execute(busca, (filtrado,))
    resultado = cursor.fetchall()
    conn.close()
    return resultado

def deletar_registro_por_id(id):
    conn=conect()
    cursor=conn.cursor()
    cursor.execute(
    ' delete from pedidos where id =?',(id,)
    )
    conn.commit()
    conn.close()
def limpar_tabela():
    conn=conect()
    cursor=conn.cursor()
    cursor.execute(
    ' delete from pedidos;'
    )
    conn.commit()
    conn.close()


def menu_pedidos():
    while True:
        print("\nMenu:")
        print("1. Criar pedido")
        print("2. atualizar tabela")
        print("3. ver dados")
        print("4. criar tabela ")
        print("5. filtrar tabela ")
        print("6. deletar registro unico ")
        print("7. limpar registros totais ")


        print("s. encerrar programa ")
        opcao = input("didite uma opção")
        if opcao == '1':
            name = input('Nome do pedido(sem extensão): ')
            n_pedido = input('Numero do pedido: ')
            data_origin = datetime.now().strftime("%d %m %Y %H:%M:%S")
            cnpj = input('Numero CNPJ contratante(formato 00.000.000/0001-00):  ')
            contratante = input("Contratante: ")
            endereco_contratante = input("Endereço do Contratante: ")
            n_contrato = input('Contrato Nº: ')
            Ata_RP = input('Numero de Ata RP(se houver): ')
            empenho = input('Numero de Empenho: ')
            o_f = input('Numero de Ordem de Fornecimento(se houver): ')
            data_recebimento_empenho = input('Data do Recebimento do Empenho(formato: DD/MM/AAAA): ')
            nf = input('Numero de Nota fiscal(quando emitida)º: ')
            contato = input('Nome do contato: ')
            telefone = input('Numero de telefone do contato: ')
            email = input('Email do contato: ')
            objeto = input('Objeto do pedido: ')
            data_hora_entrega = input('Data e hora de entrega do pedido(formato DD/MM/AAAA 00:00): ')
            local_entrega = input('Local da entrega/prestação ')
            U_F = input('Unidade de fornecimento: ')
            qtde = input('Quantidade em função da unidade de fornecimento: ')
            obs = input('Observações/ instruções para o pedido(tamanho <= 1800 caracteres): ')
            coord = input('Coordenadores responsaveis pela execução: ')

            criar_pedido(name, n_pedido, data_origin, cnpj, contratante, endereco_contratante,
                         n_contrato, Ata_RP, empenho, o_f, data_recebimento_empenho, nf, contato,
                         telefone, email, objeto, data_hora_entrega, local_entrega, U_F, qtde, obs, coord)
        elif opcao == '2':
            name = input('Nome do pedido (sem extensão) para atualizar: ')
            filename = f'{name}.xlsx'
            if os.path.exists(filename):
                print('Pedido localizado')
                id = int(input('ID do pedido que deseja atualizar: '))  # Solicita o ID
                newn_pedido = input('Novo Número do pedido: ')
                data_att = datetime.now().strftime("%d/%m/%Y %H:%M:%S")  # Data de atualização
                newcnpj = input('Novo Número CNPJ contratante (formato 00.000.000/0001-00): ')
                newcontratante = input("Novo Contratante: ")
                newendereco_contratante = input("Novo Endereço do Contratante: ")
                newn_contrato = input('Novo Contrato Nº: ')
                newAta_RP = input('Novo Número de Ata RP (se houver): ')
                newempenho = input('Novo Número de Empenho: ')
                newo_f = input('Novo Número de Ordem de Fornecimento (se houver): ')
                newdata_recebimento_empenho = input('Nova Data do Recebimento do Empenho (formato: DD/MM/AAAA): ')
                nf = input('Novo Número de Nota fiscal (quando emitida): ')
                newcontato = input('Novo Nome do contato: ')
                newtelefone = input('Novo Número de telefone do contato: ')
                newemail = input('Novo Email do contato: ')
                newobjeto = input('Novo Objeto do pedido: ')
                newdata_hora_entrega = input('Nova Data e hora de entrega do pedido (formato DD/MM/AAAA 00:00): ')
                newlocal_entrega = input('Novo Local da entrega/prestação: ')
                newU_F = input('Nova Unidade de fornecimento: ')
                newqtde = input('Nova Quantidade em função da unidade de fornecimento: ')
                newobs = input('Novas Observações/instruções para o pedido (tamanho <= 1800 caracteres): ')
                newcoord = input('Novos Coordenadores responsáveis pela execução: ')
                novo_valor=float('Digite o valor atualizado para o pedido:')

                atualizar_pedido(name, newn_pedido, data_att, id, newcnpj, newcontratante,
                                 newendereco_contratante, newn_contrato, newAta_RP, newempenho, newo_f,
                                 newdata_recebimento_empenho, nf, newcontato, newtelefone, newemail,
                                 newobjeto, newdata_hora_entrega, newlocal_entrega, newU_F, newqtde,
                                 newobs, newcoord)
            else:
                print('arquivo nao localizado')
        elif opcao == '3':
            pedidos = ver_dados()
            if isinstance(pedidos, list):
                for item in pedidos:
                    print(f'Pedido numero:{item[1]}, contrato:{item[2]}, empenho:{item[3]}, objeto:{item[4]}, '
                          f' data de entrega:{item[5]}, coordenador;{item[6]}, data de criação/atualização:{item[7]}'
                          f' status{item[8]}, valor R$ {item[9]:.2f}')
            else:
                item=pedidos
                print(f'Pedido numero:{item[1]}, contrato:{item[2]}, empenho:{item[3]}, objeto:{item[4]}, '
                      f' data de entrega:{item[5]}, coordenador;{item[6]}, data de criação/atualização:{item[7]}'
                      f' status{item[8]},valor R$ {item[9]:.2f}')
        elif opcao == '4':
            criar_tabela()
        elif opcao=='5':
            filtro=input('De onde  deseja buscar:')
            filtrado=input(" o que deseja buscar")
            resultado=filtro_db(filtro, filtrado)
            for pedido in resultado:
                print(f'Pedido número: {pedido[1]}, contrato: {pedido[2]}, empenho: {pedido[3]}, objeto: {pedido[4]}, '
                      f'data de entrega: {pedido[5]}, coordenador: {pedido[6]}, '
                      f'data de criação/atualização: {pedido[7]}, status: {pedido[8]}')
        elif opcao=='6':
            id = int(input('digite o id do usuario a ser deletato: '))
            deletar_registro_por_id(id)
            print(f"Usuário {id} deletado com sucesso.")
        elif opcao=='7':
            confirmacao = input('deseja realme nte limpar todos os registros da tabela: 1 para  sim')
            if confirmacao=='1':
                limpar_tabela()
            print(f"tabela limpa com sucesso.")

        elif opcao=='s':
            break
        else:
            print("digite uma opção valida")


if __name__=='__main__':
    menu_pedidos()
