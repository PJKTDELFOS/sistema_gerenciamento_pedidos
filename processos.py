import sqlite3
from openpyxl import workbook,load_workbook

def con():
    return sqlite3.connect("processos.db")


def tabela():
    conn=con()
    cursor=conn.cursor()
    cursor.execute( '''
    create table if not exists processos (
    ID integer primary key autoincrement,
    MODALIDADE text not null,
    CONTRATANTE text not null,
    NUMERO text not null,
    DATA_LIMITE text not null,
    STATUS text not null
    )
    '''
    )
    conn.commit()
    conn.close()

def inserir_dados(MODALIDADE,CONTRATANTE,NUMERO,DATA_LIMITE,STATUS):
    conn = con()
    cursor = conn.cursor()
    cursor.execute(
            'insert into processos (MODALIDADE,CONTRATANTE,NUMERO,DATA_LIMITE,STATUS)'
            'values(?,?,?,?,?)',(MODALIDADE,CONTRATANTE,NUMERO,DATA_LIMITE,STATUS)

                   )
    conn.commit()
    workbook=load_workbook('processos.xlsx')
    sheet=workbook['processos']
    next_row=sheet.max_row+1
    last_id=None

    for row in range(2,sheet.max_row+1):
        cell_value=sheet.cell(row=row,column=1).value
        if isinstance(cell_value,int):
            last_id=cell_value

    if last_id is None:
        id=1
    else:
        id=last_id+1
    sheet.cell(row=next_row, column=1).value = id
    sheet.cell(row=next_row, column=2).value = MODALIDADE
    sheet.cell(row=next_row, column=3).value = CONTRATANTE
    sheet.cell(row=next_row, column=4).value = NUMERO
    sheet.cell(row=next_row, column=5).value = DATA_LIMITE
    sheet.cell(row=next_row, column=6).value = STATUS
    workbook.save('processos.xlsx')
    conn.close()

def ver_dados():
    conn = con()
    cursor = conn.cursor()
    cursor.execute("select * from processos")
    resultado = cursor.fetchall()
    conn.close()
    return resultado


def att_processos(id,modalidade,contratante,numero,data_limite,status):
    conn = con()
    cursor = conn.cursor()
    cursor.execute(
        ' update processos set MODALIDADE = ?, CONTRATANTE = ?, NUMERO = ?, DATA_LIMITE = ?, STATUS = ?, '
        'where id = ?', (modalidade,contratante,numero,data_limite,status, id)
    )
    conn.commit()
    workbook = load_workbook('rh.xlsx')
    sheet = workbook['processos']
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
        if row[0].value == id:
            sheet.cell(row=row[0].row, column=2).value = modalidade
            sheet.cell(row=row[0].row, column=3).value = contratante
            sheet.cell(row=row[0].row, column=4).value = numero
            sheet.cell(row=row[0].row, column=5).value = data_limite
            sheet.cell(row=row[0].row, column=6).value = status
            break
            workbook.save('processos.xlsx')
    conn.close()

def deletar_usuario(id):
    conn = con()
    cursor = conn.cursor()
    cursor.execute(
        ' delete from processos where id= ?',(id,)
    )
    conn.commit()
    workbook = load_workbook('processos.xlsx')
    sheet = workbook['processos']
    id_buscado=id
    coluna_id=1
    for row in range(2,sheet.max_row+1):
        if sheet.cell(row=row,column=coluna_id).value==id_buscado:
            sheet.delete_rows(row)
            break
    workbook.save('processos.xlsx')
    conn.close()

def limpar_banco():
    conn = con()
    cursor = conn.cursor()
    cursor.execute(
        ' delete from processos;'
    )

    conn.commit()
    workbook = load_workbook('processos.xlsx')
    sheet = workbook['processos']
    sheet.delete_rows(2, sheet.max_row)
    workbook.save('processos')
    conn.close()



def menu_processos():
    while True:
        print("\nMenu:")
        print("1. Criar tabela")
        print("2. Inserir Processo")
        print("3. Ver dados")
        print("4. Atualizar processo")
        print("5. Deletar processo")
        print("6. Limpar tabela")
        print("7. Sair")
        opcao=input("didite uma opção")
        if opcao=='1':
            tabela()
            print('tabela criada ou ja existente')
        elif opcao=='2':
            while True:
                MODALIDADE=input('Modalidade:')
                CONTRATANTE=input("Contratante")
                NUMERO=input("Numero do processo:")
                DATA_LIMITE=input('Data LImite::')
                STATUS=input('Status:')
                inserir_dados(MODALIDADE,CONTRATANTE,NUMERO,DATA_LIMITE,STATUS)
                continuar = input('deseja continuar adicionando:').strip().lower()
                if continuar=='s':
                    continue
                else:
                    break
        elif opcao=='3':
            contrato=ver_dados()
            for item in contrato:
                print(f'ID: {item[0]}, '
                      f'MODALIDADE: {item[1]}, '
                      f'CONTRATANTE: {item[2]}, '
                      f'NUMERO: {item[3]}, '
                      f'DATA_LIMITE: {item[4]}, '
                      f'STATUS: {item[5]}, '
                      )
        elif opcao=='4':
            while True:
                print('NAO DEIXAR CAMPO VAZIO, VAI APAGAR O QUE JA ESTA ESCRITO!!!!')
                id=int(input("digite o id do usuario a ser modificado:"))
                modalidade=input('Modalidade: ')
                contratante=input("Contratante:")
                numero=input('digite telefone atual: ')
                data_limite=input('digite endereço atual: ')
                status=input('digite cargo atual: ')
                att_processos(id,modalidade,contratante,numero,data_limite,status)
                print(f"Usuário {id} atualizado com sucesso.")
                continuar = input('deseja continuar adicionando:').strip().lower()
                if continuar == 's':
                    continue
                else:
                    break
        elif opcao=='5':
            id=int(input('digite o id do usuario a ser deletato: '))
            deletar_usuario(id)
            print(f"Usuário {id} deletado com sucesso.")
        elif opcao == '6':
            limpar_banco()
        elif opcao=='7':
            break

        else:
            print("digite uma opção valida")

if __name__=='__main__':
    menu_processos()





