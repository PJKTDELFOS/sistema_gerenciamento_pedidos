import sqlite3
from openpyxl import load_workbook
def conectar():
    return sqlite3.connect("recursos_humanos.db")
def criar_tabela():
    conn=conectar()
    cursor=conn.cursor()
    cursor.execute('''
    create table if not exists usuarios (
    id integer primary key autoincrement,
    nome text not null,
    idade integer not null,
    telefone text not null,
    endereco text not null,
    cargo text not null,
    salario float not null
    )
    ''')
    conn.commit()
    conn.close()
def criar_usuario(nome,idade,telefone,endereco,cargo,salario):
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute(' insert into usuarios (nome,idade,telefone,endereco,cargo,salario)'
                   ' values (?,?,?,?,?,?)',(nome,idade,telefone,endereco,cargo,salario))
    conn.commit()
    workbook=load_workbook('rh.xlsx')
    sheet=workbook['sheet']
    next_row=sheet.max_row+1
    last_id=None


    for row in range(2,sheet.max_row+1):
       cell_value=sheet.cell(row=row,column=1).value
       if isinstance(cell_value,int):
           last_id=cell_value
    if last_id is None:
        id=1
    else:
        id=last_id+1
    sheet.cell(row=next_row, column=1).value = id
    sheet.cell(row=next_row, column=2).value = nome
    sheet.cell(row=next_row, column=3).value = idade
    sheet.cell(row=next_row, column=4).value = telefone
    sheet.cell(row=next_row, column=5).value = endereco
    sheet.cell(row=next_row, column=6).value = cargo
    sheet.cell(row=next_row, column=7).value = salario
    workbook.save('rh.xlsx')
    conn.close()
def ver_dados():
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute("select * from usuarios")
    resultado=cursor.fetchall()
    conn.close()
    return resultado

def atualizar_usuarios(id,novo_nome,nova_idade,novo_telefone,novo_endereco,novo_cargo,novo_salario):
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute(
        ' update usuarios set nome = ?, idade = ?, telefone = ?, endereco = ?, cargo = ?, '
        'salario = ? where id = ?',( novo_nome,nova_idade,
                                                novo_telefone,novo_endereco,
                                                novo_cargo,novo_salario,id)
    )
    conn.commit()
    workbook=load_workbook('rh.xlsx')
    sheet=workbook['sheet']
    for row in sheet.iter_rows(min_row=2,max_row=sheet.max_row,min_col=1,max_col=1):
        if row[0].value==id:
            sheet.cell(row=row[0].row,column=2).value=novo_nome
            sheet.cell(row=row[0].row, column=3).value = nova_idade
            sheet.cell(row=row[0].row, column=4).value = novo_telefone
            sheet.cell(row=row[0].row, column=5).value = novo_endereco
            sheet.cell(row=row[0].row, column=6).value = novo_cargo
            sheet.cell(row=row[0].row, column=7).value = novo_salario
            break
            workbook.save('rh.xlsx')
    conn.close()
def deletar_usuario(id):
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute(
        ' delete from usuarios where id= ?',(id,)
    )
    conn.commit()
    workbook = load_workbook('rh.xlsx')
    sheet = workbook['sheet']
    id_buscado=id
    coluna_id=1
    for row in range(2,sheet.max_row+1):
        if sheet.cell(row=row,column=coluna_id).value==id_buscado:
            sheet.delete_rows(row)
            break
    workbook.save('rh.xlsx')
    conn.close()
def limpar_banco():
    conn = conectar()
    cursor = conn.cursor()
    cursor.execute(
        ' delete from usuarios;'
    )

    conn.commit()
    workbook = load_workbook('rh.xlsx')
    sheet = workbook['sheet']
    sheet.delete_rows(2, sheet.max_row)
    workbook.save('rh.xlsx')
    conn.close()

def menu_rh():
    while True:
        print("\nMenu:")
        print("1. Criar tabela")
        print("2. Criar usuário")
        print("3. Ver dados")
        print("4. Atualizar usuário")
        print("5. Deletar usuário")
        print("6. Limpar tabela")
        print("7. Sair")
        opcao=input("didite uma opção")
        if opcao=='1':
            criar_tabela()
            print('tabela criada ou ja existente')
        elif opcao=='2':
            while True:
                nome=input('digite seu nome:')
                idade=int(input("digite sua idade"))
                telefone=input("digite seu telefone:")
                endereco=input('didigite seu endereço completo:')
                cargo=input('digite o cargo:')
                salario=float(input("digite o salario do funcionario: "))

                criar_usuario(nome,idade,telefone,endereco, cargo, salario)
                continuar = input('deseja continuar adicionando:').strip().lower()
                if continuar=='s':
                    continue
                else:
                    break
        elif opcao=='3':
            usuarios=ver_dados()
            for usuario in usuarios:
                print(f'ID: {usuario[0]}, '
                      f'Nome: {usuario[1]}, '
                      f'Idade: {usuario[2]}, '
                      f'Telefone: {usuario[3]}, '
                      f'Endereço: {usuario[4]}, '
                      f'Cargo: {usuario[5]}, '
                      f'Salário: {usuario[6]:.2f}')
        elif opcao=='4':
            while True:
                id=int(input("digite o id do usuario a ser modificado:"))
                novo_nome=input('digite nome atual: ')
                nova_idade=int(input("digite a idade atualizada:"))
                novo_telefone=input('digite telefone atual: ')
                novo_endereco=input('digite endereço atual: ')
                novo_cargo=input('digite cargo atual: ')
                novo_salario=float(input("digite o salario atualizadodo funcionario: "))
                atualizar_usuarios(id,novo_nome,nova_idade,novo_telefone, novo_endereco, novo_cargo, novo_salario)
                print(f"Usuário {id} atualizado com sucesso.")
                continuar = input('deseja continuar adicionando:').strip().lower()
                if continuar == 's':
                    continue
                else:
                    break
        elif opcao=='5':
            id=int(input('digite o id do usuario a ser deletato: '))
            deletar_usuario(id)
            print(f"Usuário {id} deletado com sucesso.")
        elif opcao == '6':
            limpar_banco()
        elif opcao=='7':
            break

        else:
            print("digite uma opção valida")


if __name__=='__main__':
    menu_rh()
